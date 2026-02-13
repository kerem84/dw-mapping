#!/usr/bin/env python3
"""
Mapping Excel uretici yardimci modul.
Claude tarafindan olusturulan mapping satirlarini 15 sutunluk standart formatta Excel'e yazar.

Kullanim:
  python generate_mapping.py --input mapping_data.json --output Attribute_Level_Mapping.xlsx

Girdi JSON formati:
[
  {
    "source_system": "PostgreSQL",
    "source_system_short": "KEY",
    "source_db": "keyuygulama",
    "source_schema": "public",
    "source_table": "kazaincelemeraporu",
    "master_detail": "Master",
    "target_schema": "DWH",
    "source_attribute": "kazatarihi",
    "target_physical_name": "f_kaza",
    "target_logical_name": "Kaza Master",
    "target_attribute": "kaza_tarihi",
    "schema_code": "KEY_KS001",
    "modul": "KEY",
    "kullanim_senaryosu": "KS_001",
    "senaryo_adimi": "Kaza-Risk Korelasyon"
  }
]
"""

import json
import argparse
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Sablon sutun sirasi
TEMPLATE_COLUMNS = [
    "Source System",
    "Source System.1",
    "Source Db",
    "Source Schema",
    "Source Table",
    "Master/Detail",
    "Target Schema",
    "Source Attribute",
    "Target Physical Name",
    "Target Logical Name",
    "Target Attribute",
    "Schema Code",
    "Modul",
    "Kullanim Senaryosu",
    "Senaryo Adimi",
]

# JSON key -> sutun adi eslesmesi
JSON_KEY_MAP = {
    "source_system": "Source System",
    "source_system_short": "Source System.1",
    "source_db": "Source Db",
    "source_schema": "Source Schema",
    "source_table": "Source Table",
    "master_detail": "Master/Detail",
    "target_schema": "Target Schema",
    "source_attribute": "Source Attribute",
    "target_physical_name": "Target Physical Name",
    "target_logical_name": "Target Logical Name",
    "target_attribute": "Target Attribute",
    "schema_code": "Schema Code",
    "modul": "Modul",
    "kullanim_senaryosu": "Kullanim Senaryosu",
    "senaryo_adimi": "Senaryo Adimi",
}

# Renk kodlari
FACT_FILL = PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid")  # Mavi
DIM_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Yesil
BRIDGE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Sari
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
DATA_FONT = Font(size=10, name="Calibri")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Sutun genislikleri
COLUMN_WIDTHS = {
    "Source System": 15,
    "Source System.1": 14,
    "Source Db": 18,
    "Source Schema": 14,
    "Source Table": 28,
    "Master/Detail": 14,
    "Target Schema": 14,
    "Source Attribute": 28,
    "Target Physical Name": 28,
    "Target Logical Name": 24,
    "Target Attribute": 28,
    "Schema Code": 14,
    "Modul": 10,
    "Kullanim Senaryosu": 22,
    "Senaryo Adimi": 28,
}


def get_row_fill(target_physical_name):
    """Hedef tablo adina gore renk belirle."""
    if not target_physical_name:
        return None
    name = target_physical_name.lower()
    if name.startswith("f_"):
        return FACT_FILL
    elif name.startswith("d_"):
        return DIM_FILL
    elif name.startswith("b_"):
        return BRIDGE_FILL
    return None


def generate_excel(rows, output_path):
    """Mapping satirlarini formatli Excel'e yaz."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attribute_Level_Mapping"

    # Header yaz
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Veri satirlari yaz
    for row_idx, row_data in enumerate(rows, 2):
        target_phys = row_data.get("target_physical_name", "")
        row_fill = get_row_fill(target_phys)

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
            # JSON key'i bul
            json_key = None
            for k, v in JSON_KEY_MAP.items():
                if v == col_name:
                    json_key = k
                    break

            value = row_data.get(json_key, "") if json_key else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if row_fill:
                cell.fill = row_fill

    # Sutun genislikleri
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 15)

    # Freeze pane (header sabit)
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Header satir yuksekligi
    ws.row_dimensions[1].height = 30

    wb.save(output_path)
    print(f"Mapping Excel yazildi: {output_path} ({len(rows)} satir)")

    # Istatistik
    facts = set()
    dims = set()
    bridges = set()
    for r in rows:
        tp = r.get("target_physical_name", "").lower()
        if tp.startswith("f_"):
            facts.add(tp)
        elif tp.startswith("d_"):
            dims.add(tp)
        elif tp.startswith("b_"):
            bridges.add(tp)

    print(f"Istatistik: {len(facts)} fact, {len(dims)} dim, {len(bridges)} bridge tablo")
    print(f"Toplam attribute: {len(rows)}")


def main():
    parser = argparse.ArgumentParser(description="Mapping Excel uretici")
    parser.add_argument("--input", required=True, help="Girdi JSON dosyasi")
    parser.add_argument("--output", required=True, help="Cikti Excel dosya yolu")

    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        rows = json.load(f)

    if not isinstance(rows, list):
        print("Hata: JSON dosyasi bir liste (array) icermeli.", file=sys.stderr)
        sys.exit(1)

    generate_excel(rows, args.output)
    print("Tamamlandi.")


if __name__ == "__main__":
    main()
