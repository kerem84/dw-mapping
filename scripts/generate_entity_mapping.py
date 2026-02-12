#!/usr/bin/env python3
"""
Entity Level Mapping Excel uretici.
Attribute-level mapping JSON verisinden entity-level ozet uretir.
Her benzersiz hedef tablo (Target Physical Name) icin tek satir yazar.

Kullanim:
  python generate_entity_mapping.py --input mapping_data.json --output Entity_Level_Mapping.xlsx

Girdi: generate_mapping.py ile ayni JSON formati (mapping_data.json)
Cikti: 8 sutunluk Entity_Level_Mapping.xlsx
"""

import json
import argparse
import sys
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEMPLATE_COLUMNS = [
    "Target Schema",
    "Target Physical Name",
    "Target Logical Name",
    "Source System",
    "Source Db",
    "Source Schema",
    "Source Table",
    "Modul",
]

HEADER_FILL = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, name="Calibri")
DATA_FONT = Font(size=10, name="Calibri")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMN_WIDTHS = {
    "Target Schema": 16,
    "Target Physical Name": 28,
    "Target Logical Name": 28,
    "Source System": 15,
    "Source Db": 18,
    "Source Schema": 14,
    "Source Table": 28,
    "Modul": 10,
}


def aggregate_entities(rows):
    """Attribute satirlarindan entity-level ozet cikar."""
    entities = OrderedDict()

    for row in rows:
        target_phys = row.get("target_physical_name", "")
        if not target_phys:
            continue

        if target_phys not in entities:
            entities[target_phys] = {
                "target_schema": "dwh",
                "target_physical_name": target_phys,
                "target_logical_name": row.get("target_logical_name", ""),
                "source_system": row.get("source_system", ""),
                "source_db": row.get("source_db", ""),
                "source_schema": row.get("source_schema", ""),
                "source_table": row.get("source_table", ""),
                "modul": row.get("modul", ""),
            }
        else:
            # Eger logical name bos kalmissa, Master satirindan al
            if not entities[target_phys]["target_logical_name"]:
                logical = row.get("target_logical_name", "")
                if logical:
                    entities[target_phys]["target_logical_name"] = logical

    return list(entities.values())


def generate_entity_excel(entity_rows, output_path):
    """Entity mapping satirlarini formatli Excel'e yaz."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Source-DWH"

    # JSON key -> sutun adi eslesmesi
    key_map = {
        "target_schema": "Target Schema",
        "target_physical_name": "Target Physical Name",
        "target_logical_name": "Target Logical Name",
        "source_system": "Source System",
        "source_db": "Source Db",
        "source_schema": "Source Schema",
        "source_table": "Source Table",
        "modul": "Modul",
    }
    reverse_map = {v: k for k, v in key_map.items()}

    # Header yaz
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Veri satirlari yaz
    for row_idx, row_data in enumerate(entity_rows, 2):
        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
            json_key = reverse_map.get(col_name, "")
            value = row_data.get(json_key, "") if json_key else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Sutun genislikleri
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 15)

    # Freeze pane
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Header satir yuksekligi
    ws.row_dimensions[1].height = 30

    wb.save(output_path)
    print(f"Entity Mapping Excel yazildi: {output_path} ({len(entity_rows)} entity)")


def main():
    parser = argparse.ArgumentParser(description="Entity Level Mapping Excel uretici")
    parser.add_argument("--input", required=True, help="Girdi JSON dosyasi (attribute-level mapping)")
    parser.add_argument("--output", required=True, help="Cikti Excel dosya yolu")

    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        rows = json.load(f)

    if not isinstance(rows, list):
        print("Hata: JSON dosyasi bir liste (array) icermeli.", file=sys.stderr)
        sys.exit(1)

    entity_rows = aggregate_entities(rows)
    generate_entity_excel(entity_rows, args.output)
    print("Tamamlandi.")


if __name__ == "__main__":
    main()
