#!/usr/bin/env python3
"""
Genel amacli veritabani metadata cekici.
Desteklenen DB tipleri: PostgreSQL, MSSQL
Cikti: Excel dosyasi (openpyxl + pandas)

Kullanim:
  python extract_metadata.py --type postgresql --host X --port 5432 --db mydb --user U --pass P --output metadata.xlsx
  python extract_metadata.py --type mssql --host X --port 1433 --db mydb --user U --pass P --output metadata.xlsx
  python extract_metadata.py --connection-file conn.txt --output metadata.xlsx
"""

import argparse
import sys
import re
import pandas as pd

METADATA_COLUMNS = [
    "table_schema", "table_name", "column_name", "kolon_sirasi",
    "data_type", "max_uzunluk", "numeric_precision", "numeric_scale",
    "is_nullable", "column_default", "is_primary_key", "pk_constraint",
    "is_foreign_key", "fk_constraint", "referenced_schema",
    "referenced_table", "referenced_column"
]

POSTGRESQL_SQL = """
SELECT
    c.table_schema,
    c.table_name,
    c.column_name,
    c.ordinal_position AS kolon_sirasi,
    c.data_type,
    c.character_maximum_length AS max_uzunluk,
    c.numeric_precision,
    c.numeric_scale,
    c.is_nullable,
    c.column_default,
    CASE WHEN pk.column_name IS NOT NULL THEN 'YES' ELSE 'NO' END AS is_primary_key,
    pk.constraint_name AS pk_constraint,
    CASE WHEN fk.column_name IS NOT NULL THEN 'YES' ELSE 'NO' END AS is_foreign_key,
    fk.constraint_name AS fk_constraint,
    ccu.table_schema AS referenced_schema,
    ccu.table_name AS referenced_table,
    ccu.column_name AS referenced_column
FROM information_schema.columns c
LEFT JOIN (
    SELECT kcu.table_schema, kcu.table_name, kcu.column_name, kcu.constraint_name
    FROM information_schema.key_column_usage kcu
    JOIN information_schema.table_constraints tc
        ON kcu.constraint_name = tc.constraint_name
        AND kcu.table_schema = tc.table_schema
    WHERE tc.constraint_type = 'PRIMARY KEY'
) pk ON c.table_schema = pk.table_schema
    AND c.table_name = pk.table_name
    AND c.column_name = pk.column_name
LEFT JOIN (
    SELECT kcu.table_schema, kcu.table_name, kcu.column_name, kcu.constraint_name
    FROM information_schema.key_column_usage kcu
    JOIN information_schema.table_constraints tc
        ON kcu.constraint_name = tc.constraint_name
        AND kcu.table_schema = tc.table_schema
    WHERE tc.constraint_type = 'FOREIGN KEY'
) fk ON c.table_schema = fk.table_schema
    AND c.table_name = fk.table_name
    AND c.column_name = fk.column_name
LEFT JOIN information_schema.constraint_column_usage ccu
    ON fk.constraint_name = ccu.constraint_name
    AND c.table_schema = fk.table_schema
WHERE c.table_schema NOT IN ('pg_catalog', 'information_schema')
ORDER BY c.table_schema, c.table_name, c.ordinal_position;
"""

MSSQL_SQL = """
SELECT
    s.name AS table_schema,
    t.name AS table_name,
    c.name AS column_name,
    c.column_id AS kolon_sirasi,
    ty.name AS data_type,
    c.max_length AS max_uzunluk,
    c.[precision] AS numeric_precision,
    c.scale AS numeric_scale,
    CASE WHEN c.is_nullable = 1 THEN 'YES' ELSE 'NO' END AS is_nullable,
    dc.definition AS column_default,
    CASE WHEN pk_col.column_id IS NOT NULL THEN 'YES' ELSE 'NO' END AS is_primary_key,
    pk_kc.name AS pk_constraint,
    CASE WHEN fk_col.parent_column_id IS NOT NULL THEN 'YES' ELSE 'NO' END AS is_foreign_key,
    fk_obj.name AS fk_constraint,
    ref_s.name AS referenced_schema,
    ref_t.name AS referenced_table,
    ref_c.name AS referenced_column
FROM sys.columns c
JOIN sys.tables t ON c.object_id = t.object_id
JOIN sys.schemas s ON t.schema_id = s.schema_id
JOIN sys.types ty ON c.user_type_id = ty.user_type_id
LEFT JOIN sys.default_constraints dc ON c.default_object_id = dc.object_id
LEFT JOIN (
    SELECT ic.object_id, ic.column_id, kc.name
    FROM sys.index_columns ic
    JOIN sys.indexes i ON ic.object_id = i.object_id AND ic.index_id = i.index_id
    JOIN sys.key_constraints kc ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
    WHERE i.is_primary_key = 1
) pk_col ON c.object_id = pk_col.object_id AND c.column_id = pk_col.column_id
LEFT JOIN (
    SELECT ic.object_id, ic.column_id, kc.name
    FROM sys.index_columns ic
    JOIN sys.indexes i ON ic.object_id = i.object_id AND ic.index_id = i.index_id
    JOIN sys.key_constraints kc ON i.object_id = kc.parent_object_id AND i.index_id = kc.unique_index_id
    WHERE i.is_primary_key = 1
) pk_kc ON c.object_id = pk_kc.object_id AND c.column_id = pk_kc.column_id
LEFT JOIN sys.foreign_key_columns fk_col
    ON c.object_id = fk_col.parent_object_id AND c.column_id = fk_col.parent_column_id
LEFT JOIN sys.foreign_keys fk_obj
    ON fk_col.constraint_object_id = fk_obj.object_id
LEFT JOIN sys.tables ref_t ON fk_col.referenced_object_id = ref_t.object_id
LEFT JOIN sys.schemas ref_s ON ref_t.schema_id = ref_s.schema_id
LEFT JOIN sys.columns ref_c
    ON fk_col.referenced_object_id = ref_c.object_id
    AND fk_col.referenced_column_id = ref_c.column_id
ORDER BY s.name, t.name, c.column_id;
"""


def parse_connection_file(filepath):
    """Connection string dosyasini parse et."""
    config = {}
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read().strip()

    # JDBC format: jdbc:postgresql://host:port/db
    jdbc_pg = re.match(r"jdbc:postgresql://([^:]+):(\d+)/(\S+)", content)
    if jdbc_pg:
        config["type"] = "postgresql"
        config["host"] = jdbc_pg.group(1)
        config["port"] = int(jdbc_pg.group(2))
        config["db"] = jdbc_pg.group(3)

    jdbc_ms = re.match(r"jdbc:sqlserver://([^:;]+)(?::(\d+))?.*?databaseName=(\S+)", content)
    if jdbc_ms:
        config["type"] = "mssql"
        config["host"] = jdbc_ms.group(1)
        config["port"] = int(jdbc_ms.group(2)) if jdbc_ms.group(2) else 1433
        config["db"] = jdbc_ms.group(3)

    # Satir bazli format: key=value
    for line in content.splitlines():
        line = line.strip()
        if "=" in line:
            key, val = line.split("=", 1)
            key = key.strip().lower()
            val = val.strip()
            if key in ("type", "dbtype", "db_type"):
                config["type"] = val.lower()
            elif key in ("host", "server"):
                config["host"] = val
            elif key in ("port",):
                config["port"] = int(val)
            elif key in ("db", "database", "dbname"):
                config["db"] = val
            elif key in ("user", "username"):
                config["user"] = val
            elif key in ("pass", "password"):
                config["pass"] = val

    return config


def connect_postgresql(config):
    """PostgreSQL baglantisi olustur."""
    import psycopg2
    return psycopg2.connect(
        host=config["host"],
        port=config.get("port", 5432),
        dbname=config["db"],
        user=config["user"],
        password=config["pass"],
    )


def connect_mssql(config):
    """MSSQL baglantisi olustur."""
    import pymssql
    return pymssql.connect(
        server=config["host"],
        port=config.get("port", 1433),
        database=config["db"],
        user=config["user"],
        password=config["pass"],
    )


def fetch_metadata(config):
    """Belirtilen DB tipine gore metadata cek, DataFrame dondur."""
    db_type = config["type"].lower()

    if db_type == "postgresql":
        conn = connect_postgresql(config)
        sql = POSTGRESQL_SQL
    elif db_type == "mssql":
        conn = connect_mssql(config)
        sql = MSSQL_SQL
    else:
        raise ValueError(f"Desteklenmeyen DB tipi: {db_type}. Desteklenen: postgresql, mssql")

    try:
        df = pd.read_sql(sql, conn)
        df.columns = METADATA_COLUMNS[: len(df.columns)]
        return df
    finally:
        conn.close()


def write_excel(df, output_path):
    """DataFrame'i formatlı Excel dosyasina yaz."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Metadata")
        ws = writer.sheets["Metadata"]

        # Header stili
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Veri hucreleri border
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).border = thin_border

        # Sutun genislikleri
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df.iloc[:, col_idx - 1].astype(str).str.len().max())
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 30)

        # Freeze pane ve auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"Metadata yazildi: {output_path} ({len(df)} satir)")


def main():
    parser = argparse.ArgumentParser(description="Veritabani metadata cekici")
    parser.add_argument("--type", choices=["postgresql", "mssql"], help="DB tipi")
    parser.add_argument("--host", help="DB host")
    parser.add_argument("--port", type=int, help="DB port")
    parser.add_argument("--db", help="Veritabani adi")
    parser.add_argument("--user", help="Kullanici adi")
    parser.add_argument("--pass", dest="password", help="Sifre")
    parser.add_argument("--connection-file", help="Connection string dosyasi")
    parser.add_argument("--output", required=True, help="Cikti Excel dosya yolu")

    args = parser.parse_args()

    if args.connection_file:
        config = parse_connection_file(args.connection_file)
        # CLI parametreleri override
        if args.type:
            config["type"] = args.type
        if args.user:
            config["user"] = args.user
        if args.password:
            config["pass"] = args.password
    else:
        if not all([args.type, args.host, args.db, args.user, args.password]):
            parser.error("--type, --host, --db, --user, --pass gerekli (veya --connection-file kullanin)")
        config = {
            "type": args.type,
            "host": args.host,
            "port": args.port,
            "db": args.db,
            "user": args.user,
            "pass": args.password,
        }

    print(f"Baglaniyor: {config['type']}://{config['host']}:{config.get('port', 'default')}/{config['db']}")
    df = fetch_metadata(config)
    write_excel(df, args.output)
    print("Tamamlandi.")


if __name__ == "__main__":
    main()
